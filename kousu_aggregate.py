import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import openpyxl
from io import BytesIO
import os
import re
import datetime

# フォント自動検出
font_path_candidates = [
    "C:/Windows/Fonts/meiryo.ttc",
    "C:/Windows/Fonts/YuGothM.ttc",
]
font_path = None
for p in font_path_candidates:
    if os.path.isfile(p):
        font_path = p
        break
prop = fm.FontProperties(fname=font_path) if font_path else None

# グローバルにmatplotlibの日本語フォントを設定
if prop:
    plt.rcParams['font.family'] = prop.get_name()

st.set_page_config(layout="wide")
st.title('工数集計・可視化アプリ')

staff_to_branch = {
    "a.kani":      "東京",
    "a.murai":     "東京",
    "d.tajima":    "東京",
    "h.meguro":    "東京",
    "h.obata":     "大阪",
    "h.sato":      "東京",
    "k.fujita":    "大阪",
    "k.horio":     "大阪",
    "k.muraoka":   "大阪",
    "k.usami":     "東京",
    "m.maekawa":   "大阪",
    "m.moriguchi": "大阪",
    "m.okawa":     "大阪",
    "r.tanaka":    "大阪",
    "s.tawada":    "大阪",
    "y.hara":      "大阪",
    "y.nakai":     "東京",
    "y.nishitani": "大阪",
    "yuki.sato":   "東京",
}

work_content_list = [
    "移動", "納品試運転", "制御部更新", "点検", "年間保守", "訪問修理", "引取修理", "改造", "移設",
    "お客様対応(保証期限内)", "お客様対応(受注でない)", "社内サポート", "修繕", "貿易管理", "庶務",
    "教育", "標準化", "検査", "組立", "手配", "設計", "その他"
]

def extract_number(val):
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        m = re.match(r'^(\d+)', val)
        if m:
            return int(m.group(1))
    return pd.NA

def convert_month(month_raw):
    if month_raw is None:
        return ""
    if isinstance(month_raw, datetime.datetime):
        return month_raw.strftime("%Y_%m")
    if isinstance(month_raw, str):
        try:
            month_dt = pd.to_datetime(month_raw)
            return month_dt.strftime("%Y_%m")
        except Exception:
            return str(month_raw)
    if isinstance(month_raw, (int, float)):
        try:
            month_dt = pd.to_datetime('1899-12-30') + pd.to_timedelta(float(month_raw), unit='D')
            return month_dt.strftime("%Y_%m")
        except Exception:
            return str(month_raw)
    return str(month_raw)

# --- サイドバーで保存済みCSVデータのアップロード ---
st.sidebar.markdown("---")
st.sidebar.markdown("### 保存したCSVデータを再アップロード（積み上げ用）")
uploaded_csv = st.sidebar.file_uploader("保存済みCSVファイルを選択", type="csv", key="saved_csv")
df_saved = None
if uploaded_csv:
    df_saved = pd.read_csv(uploaded_csv)

uploaded_files = st.file_uploader(
    'Upload Excel files (multiple allowed)',
    type=['xlsx', 'xlsm'],
    accept_multiple_files=True
)

all_data = []

if uploaded_files:
    for up in uploaded_files:
        file_bytes = BytesIO(up.read())
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
        ws = wb["①個人記入欄"]
        month_raw = ws["H2"].value
        month = convert_month(month_raw)
        staff = ws["F1"].value
        branch = staff_to_branch.get(staff, "")

        file_bytes.seek(0)
        df = pd.read_excel(
            file_bytes,
            sheet_name="①個人記入欄",
            header=None,
            skiprows=6,
            usecols=[3, 4, 7, 39],
            nrows=130
        )
        df.columns = ["作業分類", "作業内容", "作業分類元", "工数 [h]"]

        cond = pd.Series([True]*len(df))
        if len(df) > 46:
            cond.iloc[46:] = df.loc[46:, "作業分類"].notna() & (df.loc[46:, "作業分類"].astype(str).str.strip() != "")
        cond = cond & ~(
            df["作業分類"].astype(str).str.startswith("参考行", na=False) |
            df["作業内容"].astype(str).str.startswith("参考行", na=False) |
            df["作業分類元"].astype(str).str.startswith("参考行", na=False) |
            df["工数 [h]"].astype(str).str.startswith("参考行", na=False)
        )

        df_valid = df[cond].copy()
        df_valid["工数 [h]"] = df_valid["工数 [h]"].apply(extract_number)
        df_valid["工数 [h]"] = pd.to_numeric(df_valid["工数 [h]"], errors="coerce")
        df_valid["作業内容_分類"] = df_valid["作業内容"].apply(
            lambda x: x if x in work_content_list else "その他"
        )
        df_valid["月"] = month
        df_valid["担当者"] = staff
        df_valid["支店"] = branch

        all_data.append(df_valid)

    df_all = pd.concat(all_data, ignore_index=True)
else:
    df_all = pd.DataFrame()  # 空データフレーム

# --- 積み上げ（結合）＆重複排除 ---
if df_saved is not None and not df_all.empty:
    df_all = pd.concat([df_saved, df_all], ignore_index=True)
    df_all = df_all.drop_duplicates()
elif df_saved is not None:
    df_all = df_saved

if not df_all.empty:
    st.subheader("抽出・除外済みデータ（積み上げ・重複排除済み）")
    st.dataframe(df_all)

    # --- サイドバー・チェックボックスフィルター ---
    st.sidebar.header("Filter Settings")

    month_list = sorted(df_all["月"].dropna().unique())
    staff_list = sorted(df_all["担当者"].dropna().unique())
    branch_list = sorted(df_all["支店"].dropna().unique())

    selected_months = []
    st.sidebar.markdown("#### 月でフィルター")
    for val in month_list:
        if st.sidebar.checkbox(val, value=True, key=f"month_{val}"):
            selected_months.append(val)

    selected_staffs = []
    st.sidebar.markdown("#### 担当者でフィルター")
    for val in staff_list:
        if st.sidebar.checkbox(val, value=True, key=f"staff_{val}"):
            selected_staffs.append(val)

    selected_branches = []
    st.sidebar.markdown("#### 支店でフィルター")
    for val in branch_list:
        if st.sidebar.checkbox(val, value=True, key=f"branch_{val}"):
            selected_branches.append(val)

    work_content_sum = df_all.groupby("作業内容_分類")["工数 [h]"].sum().to_dict()
    selected_workcontent = []
    st.sidebar.markdown("#### 作業内容でフィルター")
    for name in work_content_list:
        default_check = bool(work_content_sum.get(name, 0))
        checked = st.sidebar.checkbox(name, value=default_check, key=f"work_{name}")
        if checked:
            selected_workcontent.append(name)

    filtered = df_all[
        (df_all["月"].isin(selected_months)) &
        (df_all["担当者"].isin(selected_staffs)) &
        (df_all["支店"].isin(selected_branches)) &
        (df_all["作業内容_分類"].isin(selected_workcontent))
    ]

    st.subheader("フィルター後データ")
    st.dataframe(filtered)

    # --- ダウンロードボタン追加 ---
    csv = filtered.to_csv(index=False, encoding="utf-8-sig")
    st.download_button(
        label="このフィルター後データをCSVでダウンロード",
        data=csv,
        file_name="filtered_data.csv",
        mime='text/csv'
    )

    # --- 横棒グラフ＆円グラフ 横並び表示 ---
    st.subheader("作業内容別 工数 [h] グラフ")

    bar_data = filtered.groupby("作業内容_分類")["工数 [h]"].sum().reindex(work_content_list).fillna(0)
    pie_data = bar_data[bar_data > 0]

    col1, col2 = st.columns([2, 1])  # 横棒:円グラフ=2:1の割合

    with col1:
        st.markdown("#### 横棒グラフ")
        if bar_data.sum() > 0:
            fig, ax = plt.subplots(figsize=(8, 6))
            bar_data.plot.barh(ax=ax, color="#4a90e2")
            if prop:
                ax.set_xlabel("工数 [h]", fontproperties=prop)
                ax.set_ylabel("作業内容", fontproperties=prop)
                for label in ax.get_yticklabels():
                    label.set_fontproperties(prop)
                for label in ax.get_xticklabels():
                    label.set_fontproperties(prop)
            st.pyplot(fig, use_container_width=True)
        else:
            st.info("工数 [h] がゼロのためグラフはありません。")

    with col2:
        st.markdown("#### 円グラフ")
        if not pie_data.empty:
            fig2, ax2 = plt.subplots(figsize=(4, 4))
            labels = [str(idx) for idx in pie_data.index]
            wedges, texts, autotexts = ax2.pie(
                pie_data,
                labels=labels,
                autopct='%1.1f%%',
                counterclock=False,
                startangle=90,
                textprops={'fontproperties': prop} if prop else None
            )
            if prop:
                for text in texts + autotexts:
                    text.set_fontproperties(prop)
            ax2.set_ylabel("")
            st.pyplot(fig2, use_container_width=True)
        else:
            st.info("工数 [h] がゼロのため円グラフはありません。")

    # === 横並びで解析しやすいグラフ: 作業内容⇔担当者クロス分析 ===
    st.subheader("作業内容・担当者ごと工数 [h] 横並びグラフ")

    col_wc, col_staff = st.columns(2)

    with col_wc:
        st.markdown("#### 作業内容ごとの担当者別工数（横棒グラフ）")
        selected_wc = st.selectbox("作業内容を選択してください", work_content_list, key="wc_bar")
        df_wc = filtered[filtered["作業内容_分類"] == selected_wc]
        bar_wc_data = df_wc.groupby("担当者")["工数 [h]"].sum().sort_values()
        if not bar_wc_data.empty:
            fig3, ax3 = plt.subplots(figsize=(6, 4))
            bar_wc_data.plot.barh(ax=ax3, color="#e67e22")
            if prop:
                ax3.set_xlabel("工数 [h]", fontproperties=prop)
                ax3.set_ylabel("担当者", fontproperties=prop)
                for label in ax3.get_yticklabels():
                    label.set_fontproperties(prop)
                for label in ax3.get_xticklabels():
                    label.set_fontproperties(prop)
            st.pyplot(fig3, use_container_width=True)
        else:
            st.info("選択した作業内容のデータがありません。")

    with col_staff:
        st.markdown("#### 担当者ごとの作業内容別工数（横棒グラフ）")
        staff_options = filtered["担当者"].dropna().unique()
        if len(staff_options) > 0:
            selected_staff = st.selectbox("担当者を選択してください", staff_options, key="staff_bar")
            df_staff = filtered[filtered["担当者"] == selected_staff]
            bar_staff_data = df_staff.groupby("作業内容_分類")["工数 [h]"].sum().reindex(work_content_list).fillna(0)
            bar_staff_data = bar_staff_data[bar_staff_data > 0]
            if not bar_staff_data.empty:
                fig4, ax4 = plt.subplots(figsize=(6, 4))
                bar_staff_data.plot.barh(ax=ax4, color="#4a90e2")
                if prop:
                    ax4.set_xlabel("工数 [h]", fontproperties=prop)
                    ax4.set_ylabel("作業内容", fontproperties=prop)
                    for label in ax4.get_yticklabels():
                        label.set_fontproperties(prop)
                    for label in ax4.get_xticklabels():
                        label.set_fontproperties(prop)
                st.pyplot(fig4, use_container_width=True)
            else:
                st.info("選択した担当者のデータがありません。")
        else:
            st.info("担当者データがありません。")

else:
    st.info("Upload Excel files or upload CSV to start.")